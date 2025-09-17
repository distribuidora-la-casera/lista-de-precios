import re
import json
from datetime import datetime
from openpyxl import load_workbook

def was_price_changed(previous_data, grupo, articulo, new_price):
    try:
        old_price = previous_data["grupos"][grupo][articulo]["precio"]
        return 1 if str(old_price) != str(new_price) else 0
    except KeyError:
        return 0




previous_data_file = 'web/data.json'
input_file = 'upload/lista.xlsx'

previous_data = {}

# Load previous data
with open(previous_data_file, 'r') as f:
    previous_data = json.load(f)

# Read the Excel file
wb = load_workbook(input_file)
ws = wb["Lista"]

data = {"grupos": {}, "ofertas": {}, "fecha": datetime.now().strftime("%d-%m-%Y %H:%M:%S")}

for row in ws.iter_rows(min_row=1, values_only=True):
    grupo, articulo, precio = row
    if grupo not in data["grupos"]:
        data["grupos"][grupo] = {}
    print(precio)
    data["grupos"][grupo][articulo] = {"precio": str(precio), "cambio": was_price_changed(previous_data, grupo, articulo, precio)}

# Save the new data
with open(previous_data_file, 'w') as f:
    json.dump(data, f, indent=4, ensure_ascii=False)

